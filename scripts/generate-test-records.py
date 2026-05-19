#!/usr/bin/env python3
"""
Generate scripts/test-records.sql from the same SRBSA xlsx the import
script reads — produces ~25 agri_records spread across barangays,
commodities (Rice/Corn/Fishery), and statuses (active/harvested/damaged)
so the pilot can drive every form path against real-named farmers.

Run AFTER scripts/import-rsbsa.sql has been applied — the test records
reference the deterministic UUIDs that script creates (uuid5 with the
same NAMESPACE).

USAGE
  python scripts/generate-test-records.py \
      "C:/path/to/TUBO SRBSA.xlsx" scripts/test-records.sql
"""

from __future__ import annotations
import sys, uuid, random
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
import openpyxl

# MUST match the namespace in scripts/generate-rsbsa-sql.py so farmer +
# household UUIDs resolve to the rows the import script created.
NAMESPACE = uuid.UUID("c8b3f5e4-1a2d-4f6e-9b3c-7d5e8f9a0b1c")

BARANGAY_MAP = {
    "SUPO": "Supo", "POBLACION MAYABO": "Poblacion", "WAYANGAN": "Wayangan",
    "KILI": "Kili", "TIEMPO": "Tiempo", "AMTUAGAN": "Amtuagan",
    "TABACDA": "Tabacda", "ALANGTIN": "Alangtin", "DILONG": "Dilong",
    "TUBTUBA": "Tubtuba",
}
SEX_MAP = {"MALE": "Male", "FEMALE": "Female"}

# Test-record diversity matrix. Cycled across the selected farmers so we
# get coverage of every commodity × status combination.
COMMODITY_PLAN = [
    ("Rice",    "Inbred",    "CROP"),
    ("Rice",    "Hybrid",    "CROP"),
    ("Corn",    "",          "CROP"),
    ("Fishery", "Tilapia",   "FISHERY"),
    ("Rice",    "Traditional","CROP"),
    ("Fishery", "Carp",      "FISHERY"),
]
STATUS_PLAN = ["active", "active", "active", "harvested", "damaged"]


def sql_str(v) -> str:
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def title_case(s):
    return str(s).strip().title() if s else ""


def combine_name(last, first, middle):
    return " ".join(p for p in [first, middle, last] if p).strip()


def uuid5(key: str) -> str:
    return str(uuid.uuid5(NAMESPACE, key))


def main(xlsx_path: str, sql_path: str) -> int:
    src = Path(xlsx_path)
    if not src.exists():
        print(f"ERROR: {src} not found", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {
        "rsbsa":     header.index("SYSTEM_GENERATED_RSBSA_NUMBER"),
        "last":      header.index("LAST NAME"),
        "first":     header.index("FIRST NAME"),
        "middle":    header.index("MIDDLE NAME"),
        "barangay":  header.index("FARMER ADDRESS 1"),
        "sex":       header.index("SEX"),
        "farm_area": header.index("FARM AREA"),
    }

    # Dedup by RSBSA (matches import script behavior), bucket by canonical barangay.
    seen = set()
    by_barangay = defaultdict(list)
    for r in rows[1:]:
        rsbsa = r[idx["rsbsa"]]
        bgy_raw = r[idx["barangay"]]
        sex_raw = r[idx["sex"]]
        if not rsbsa or rsbsa in seen: continue
        if bgy_raw not in BARANGAY_MAP or sex_raw not in SEX_MAP: continue
        seen.add(rsbsa)
        by_barangay[BARANGAY_MAP[bgy_raw]].append(r)

    # Pick ~3 farmers per barangay (deterministic — seeded random so the
    # generated SQL doesn't churn between runs).
    rng = random.Random(20260519)
    selected = []
    PER_BGY = 3
    for bgy in sorted(by_barangay.keys()):
        pool = by_barangay[bgy][:]
        rng.shuffle(pool)
        selected.extend((bgy, r) for r in pool[:PER_BGY])

    # Spread commodity + status cyclically across the selection.
    rng.shuffle(selected)  # extra spread so per-barangay diversity isn't biased

    records = []
    for i, (bgy, r) in enumerate(selected):
        rsbsa = r[idx["rsbsa"]]
        last = title_case(r[idx["last"]])
        first = title_case(r[idx["first"]])
        middle = title_case(r[idx["middle"]])
        full_name = combine_name(last, first, middle)
        sex = SEX_MAP[r[idx["sex"]]]
        farm_area = float(r[idx["farm_area"]] or 0)

        farmer_id = uuid5(f"farmer:{rsbsa}")
        record_id = uuid5(f"test-record:{rsbsa}:{i}")

        commodity, sub_category, commodity_group = COMMODITY_PLAN[i % len(COMMODITY_PLAN)]
        status = STATUS_PLAN[i % len(STATUS_PLAN)]

        # Per-commodity field shape.
        planting = 0.0
        harvest_bags = 0.0
        dmg_pests = 0.0
        dmg_cal = 0.0
        stocking = 0.0
        harvest_fish = 0.0
        fish_loss = 0.0
        pests_diseases = "None"
        calamity = "None"

        if commodity_group == "CROP":
            # Cap planting at 80% of household farm area, but no more than 0.5 ha
            # so we never overallocate.
            planting = round(min(0.5, max(0.05, farm_area * 0.8)), 4)
            if status == "harvested":
                # ~80 bags/ha rice yield baseline
                harvest_bags = round(planting * 80 + rng.uniform(-5, 15), 2)
            elif status == "damaged":
                # Roughly 60% of planting area as pest loss (>=50% triggers damaged)
                dmg_pests = round(planting * 0.6, 4)
                pests_diseases = "Rats and stem borer"
        else:  # FISHERY
            stocking = round(rng.uniform(500, 1500))
            if status == "harvested":
                # ~85% recovery
                harvest_fish = round(stocking * rng.uniform(0.75, 0.90))
            elif status == "damaged":
                # ~70% loss
                fish_loss = round(stocking * 0.70)
                calamity = "Flooding"

        farmer_male = 1 if sex == "Male" else 0
        farmer_female = 1 if sex == "Female" else 0

        records.append({
            "id": record_id,
            "barangay": bgy,
            "commodity": commodity,
            "commodity_group": commodity_group,
            "sub_category": sub_category,
            "farmer_ids": f"ARRAY[{sql_str(farmer_id)}]::text[]",
            "farmer_names": full_name,
            "farmer_male": farmer_male,
            "farmer_female": farmer_female,
            "total_farmers": 1,
            "planting_area_hectares": planting,
            "harvesting_output_bags": harvest_bags,
            "damage_pests_hectares": dmg_pests,
            "damage_calamity_hectares": dmg_cal,
            "stocking": stocking,
            "harvesting_fishery": harvest_fish,
            "fishery_loss_pieces": fish_loss,
            "pests_diseases": pests_diseases,
            "calamity": calamity,
            "calamity_sub_category": "None",
            "remarks": f"[test record] {commodity} {status}",
            "status": status,
            "period_month": 5,
            "period_year": 2026,
        })

    # Build the SQL.
    out = []
    commodity_counts = Counter((r["commodity"], r["status"]) for r in records)
    barangay_counts = Counter(r["barangay"] for r in records)

    out.append("-- =========================================================================")
    out.append("-- Test agri_records — pilot smoke-test dataset.")
    out.append("-- =========================================================================")
    out.append("--")
    out.append(f"-- Generated:    {datetime.now().isoformat(timespec='seconds')}")
    out.append(f"-- Total records: {len(records)}")
    out.append("--")
    out.append("-- Commodity x status distribution:")
    for (c, s), n in sorted(commodity_counts.items()):
        out.append(f"--   {c:<10s} {s:<10s} {n}")
    out.append("--")
    out.append("-- Per-barangay distribution:")
    for bgy, n in sorted(barangay_counts.items()):
        out.append(f"--   {bgy:<12s} {n}")
    out.append("--")
    out.append("-- PRE-REQUISITES")
    out.append("--   * scripts/wipe-pilot-data.sql + scripts/import-rsbsa.sql have")
    out.append("--     been applied (this script references the deterministic UUIDs")
    out.append("--     of those imported farmers).")
    out.append("--")
    out.append("-- USAGE")
    out.append("--   Supabase Dashboard → SQL Editor → paste → Run.")
    out.append("--   Run AFTER the RSBSA import; otherwise the FK to farmers would")
    out.append("--   look up unknown UUIDs (the import populates them).")
    out.append("--")
    out.append("-- WHAT TO TEST AFTER")
    out.append("--   1. Records tab — see ~25 rows with mixed statuses + commodities.")
    out.append("--   2. Overview tab — KPI tiles populate.")
    out.append("--   3. Click Edit on an Active record — modify a field — close —")
    out.append("--      Discard confirm fires (Week 3.5 Part 8).")
    out.append("--   4. Click Edit on a Harvested record — numeric fields should be")
    out.append("--      locked for reporting integrity (Week 2 confirm-on-finalize).")
    out.append("--   5. Try to save the same record from two browser tabs — the")
    out.append("--      second save should show 'Someone else updated this record...'")
    out.append("--      (Week 3.5 Part 1 optimistic concurrency).")
    out.append("--   6. Activity tab — entries for each record creation.")
    out.append("--   7. Soft-delete one record — /admin/restore shows it — restore.")
    out.append("--")
    out.append("-- RE-RUNNING")
    out.append("--   ON CONFLICT (id) DO NOTHING — idempotent. UUIDs are deterministic.")
    out.append("--")
    out.append("")
    out.append("BEGIN;")
    out.append("")
    out.append("INSERT INTO public.agri_records (")
    out.append("  id, barangay, commodity, commodity_group, sub_category,")
    out.append("  farmer_ids, farmer_names, farmer_male, farmer_female, total_farmers,")
    out.append("  planting_area_hectares, harvesting_output_bags,")
    out.append("  damage_pests_hectares, damage_calamity_hectares,")
    out.append("  stocking, harvesting_fishery, fishery_loss_pieces,")
    out.append("  pests_diseases, calamity, calamity_sub_category,")
    out.append("  remarks, status, period_month, period_year")
    out.append(") VALUES")

    value_rows = []
    for r in records:
        value_rows.append(
            f"  ({sql_str(r['id'])}, {sql_str(r['barangay'])}, {sql_str(r['commodity'])}, "
            f"{sql_str(r['commodity_group'])}, {sql_str(r['sub_category'])}, "
            f"{r['farmer_ids']}, {sql_str(r['farmer_names'])}, "
            f"{r['farmer_male']}, {r['farmer_female']}, {r['total_farmers']}, "
            f"{r['planting_area_hectares']}, {r['harvesting_output_bags']}, "
            f"{r['damage_pests_hectares']}, {r['damage_calamity_hectares']}, "
            f"{r['stocking']}, {r['harvesting_fishery']}, {r['fishery_loss_pieces']}, "
            f"{sql_str(r['pests_diseases'])}, {sql_str(r['calamity'])}, "
            f"{sql_str(r['calamity_sub_category'])}, {sql_str(r['remarks'])}, "
            f"{sql_str(r['status'])}, {r['period_month']}, {r['period_year']})"
        )

    out.append(",\n".join(value_rows))
    out.append("ON CONFLICT (id) DO NOTHING;")
    out.append("")
    out.append("COMMIT;")
    out.append("")
    out.append("-- Verification:")
    out.append("--   SELECT status, count(*) FROM public.agri_records GROUP BY status ORDER BY status;")
    out.append(f"--   -- expect total: {len(records)}")
    out.append("--   SELECT barangay, count(*) FROM public.agri_records GROUP BY barangay ORDER BY barangay;")
    out.append("--   SELECT commodity, count(*) FROM public.agri_records GROUP BY commodity ORDER BY commodity;")

    Path(sql_path).write_text("\n".join(out) + "\n", encoding="utf-8")
    print(f"[ok] Wrote {sql_path}")
    print(f"     Total records: {len(records)}")
    for (c, s), n in sorted(commodity_counts.items()):
        print(f"       {c:<10s} {s:<10s} {n}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
