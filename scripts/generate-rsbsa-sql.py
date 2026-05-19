#!/usr/bin/env python3
"""
Generate scripts/import-rsbsa.sql from a SRBSA roster xlsx.

USAGE
  python scripts/generate-rsbsa-sql.py \
      "C:/path/to/TUBO SRBSA.xlsx" scripts/import-rsbsa.sql

The roster is expected to be a single sheet with the 22-column header
documented in the source xlsx (SYSTEM_GENERATED_RSBSA_NUMBER, LAST NAME,
FIRST NAME, MIDDLE NAME, SUFFIX AND EXTENSION, FARMER ADDRESS 1..3,
FARM ADDRESS 1, BIRTHDATE, SEX, CONTACT NO, 4Ps, Indegenous, PWD,
FARM AREA, AREA PLANTED, COMMODITY, SUB-COMMODITY/SEED, VOUCHER VALUE,
FUND SOURCE, SEASON).

This is a one-off pre-pilot data load. Re-runnable: the deterministic
UUIDs (uuid5 with a fixed namespace) mean re-running produces the same
SQL, so the import is idempotent against ON CONFLICT DO NOTHING.
"""

from __future__ import annotations

import sys
import uuid
from pathlib import Path
from datetime import datetime, date
from collections import Counter

import openpyxl

# Stable namespace so re-running the generator produces the same UUIDs.
# Random-but-fixed UUID; do NOT change this in future runs.
NAMESPACE = uuid.UUID("c8b3f5e4-1a2d-4f6e-9b3c-7d5e8f9a0b1c")

# FARMER ADDRESS 1 (xlsx) → canonical BARANGAYS (lib/data.ts).
BARANGAY_MAP = {
    "SUPO": "Supo",
    "POBLACION MAYABO": "Poblacion",
    "WAYANGAN": "Wayangan",
    "KILI": "Kili",
    "TIEMPO": "Tiempo",
    "AMTUAGAN": "Amtuagan",
    "TABACDA": "Tabacda",
    "ALANGTIN": "Alangtin",
    "DILONG": "Dilong",
    "TUBTUBA": "Tubtuba",
}

# SEX (xlsx) → gender (DB CHECK constraint accepts 'Male' | 'Female').
SEX_MAP = {"MALE": "Male", "FEMALE": "Female"}


def sql_escape(value: str | None) -> str:
    """Single-quote a string for safe SQL literal use."""
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def title_case_name(raw: str | None) -> str:
    """Convert 'AGWAYAS' → 'Agwayas'. Filipino name conventions are simple
    enough that Python's .title() suffices."""
    if not raw:
        return ""
    return str(raw).strip().title()


def combine_name(last: str, first: str, middle: str) -> str:
    """Match lib/name.ts combineNameParts: '{first} {middle} {last}'."""
    parts = [p for p in [first, middle, last] if p]
    return " ".join(parts).strip()


def deterministic_uuid(key: str) -> str:
    """uuid5 keeps the import idempotent across re-runs."""
    return str(uuid.uuid5(NAMESPACE, key))


def fmt_date(d) -> str:
    """datetime/date → 'YYYY-MM-DD' SQL DATE literal, or NULL."""
    if d is None:
        return "NULL"
    if isinstance(d, datetime):
        return sql_escape(d.date().isoformat())
    if isinstance(d, date):
        return sql_escape(d.isoformat())
    # Fallback: try to parse common string formats.
    s = str(d).strip()
    return sql_escape(s) if s else "NULL"


def fmt_float(v) -> str:
    """numeric → SQL double precision literal, or NULL."""
    if v is None or v == "":
        return "NULL"
    try:
        return f"{float(v):.4f}"
    except (TypeError, ValueError):
        return "NULL"


def main(xlsx_path: str, sql_path: str) -> int:
    src = Path(xlsx_path)
    if not src.exists():
        print(f"ERROR: source xlsx not found: {src}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    # Column indices — verified from header
    IDX = {
        "rsbsa":     header.index("SYSTEM_GENERATED_RSBSA_NUMBER"),
        "last":      header.index("LAST NAME"),
        "first":     header.index("FIRST NAME"),
        "middle":    header.index("MIDDLE NAME"),
        "barangay":  header.index("FARMER ADDRESS 1"),
        "birthdate": header.index("BIRTHDATE"),
        "sex":       header.index("SEX"),
        "farm_area": header.index("FARM AREA"),
    }

    # Dedupe by RSBSA, first occurrence wins. SRBSA roster has 93 duplicate
    # RSBSA numbers in the 1191-row source — they appear to be re-submissions
    # for different seasons/voucher rounds rather than distinct farmers.
    seen_rsbsa: set[str] = set()
    unique_rows: list[tuple] = []
    skipped_dupes = 0
    skipped_unmapped_barangay: list[str] = []
    skipped_unmapped_sex: list[str] = []

    for r in data_rows:
        rsbsa = r[IDX["rsbsa"]]
        if not rsbsa:
            continue
        if rsbsa in seen_rsbsa:
            skipped_dupes += 1
            continue
        bgy_raw = r[IDX["barangay"]]
        sex_raw = r[IDX["sex"]]
        if bgy_raw not in BARANGAY_MAP:
            skipped_unmapped_barangay.append(str(bgy_raw))
            continue
        if sex_raw not in SEX_MAP:
            skipped_unmapped_sex.append(str(sex_raw))
            continue
        seen_rsbsa.add(rsbsa)
        unique_rows.append(r)

    barangay_counts = Counter(BARANGAY_MAP[r[IDX["barangay"]]] for r in unique_rows)

    out_lines: list[str] = []
    out_lines.append("-- =========================================================================")
    out_lines.append("-- RSBSA import — TUBO SRBSA roster → households + farmers.")
    out_lines.append("-- =========================================================================")
    out_lines.append("--")
    out_lines.append(f"-- Generated from: {src.name}")
    out_lines.append(f"-- Generated at:   {datetime.now().isoformat(timespec='seconds')}")
    out_lines.append(f"-- Source rows:    {len(data_rows)}")
    out_lines.append(f"-- Unique imports: {len(unique_rows)} (dedupe by RSBSA, first wins)")
    out_lines.append(f"-- Skipped dupes:  {skipped_dupes}")
    out_lines.append(f"-- Skipped (bad barangay): {len(skipped_unmapped_barangay)}")
    out_lines.append(f"-- Skipped (bad sex):      {len(skipped_unmapped_sex)}")
    out_lines.append("--")
    out_lines.append("-- Per-barangay imported counts:")
    for bgy, count in sorted(barangay_counts.items()):
        out_lines.append(f"--   {bgy:<12s} {count}")
    out_lines.append("--")
    out_lines.append("-- ASSUMPTIONS")
    out_lines.append("--   * Every RSBSA-registered farmer becomes a household head of a")
    out_lines.append("--     1-person household. Multi-farmer households need to be linked")
    out_lines.append("--     manually after the import (Farmers tab → edit farmer →")
    out_lines.append("--     household). The RSBSA roster doesn't carry household membership.")
    out_lines.append("--   * Each household's farming_area_hectares = the row's FARM AREA.")
    out_lines.append("--   * UUIDs are deterministic (uuid5 with a fixed namespace) so")
    out_lines.append("--     re-running the generator on the same xlsx produces the same SQL.")
    out_lines.append("--   * Names are title-cased ('AGWAYAS' → 'Agwayas'). Suffix column is")
    out_lines.append("--     ignored — most rows have NULL and the form doesn't model it.")
    out_lines.append("--   * Birthdate, sex, RSBSA number, barangay are imported. Civil")
    out_lines.append("--     status + photo_url + contact number are NOT (no schema column")
    out_lines.append("--     for contact yet; pilot can add later).")
    out_lines.append("--")
    out_lines.append("-- USAGE")
    out_lines.append("--   1. Run scripts/wipe-pilot-data.sql FIRST. The import below assumes")
    out_lines.append("--      farmers + households tables are empty so RSBSA uniqueness isn't")
    out_lines.append("--      polluted by mock data.")
    out_lines.append("--   2. Supabase Dashboard → SQL Editor → paste this entire file → Run.")
    out_lines.append("--      The Editor's service-role credentials bypass RLS.")
    out_lines.append("--   3. Verify (queries at the bottom). Farmers count should equal the")
    out_lines.append(f"--      'Unique imports' number above ({len(unique_rows)}).")
    out_lines.append("--")
    out_lines.append("-- RE-RUNNING")
    out_lines.append("--   Each household + farmer INSERT uses ON CONFLICT (id) DO NOTHING,")
    out_lines.append("--   so re-running on a populated DB is idempotent — no duplicates, no")
    out_lines.append("--   updates. Use the wipe + re-import flow if you want a true reset.")
    out_lines.append("--")
    out_lines.append("")
    out_lines.append("BEGIN;")
    out_lines.append("")
    out_lines.append("-- ─────────────────────────────────────────────────────────────────────")
    out_lines.append(f"-- Households ({len(unique_rows)} rows)")
    out_lines.append("-- ─────────────────────────────────────────────────────────────────────")
    out_lines.append("")
    out_lines.append("INSERT INTO public.households (id, barangay, display_name, farming_area_hectares) VALUES")

    household_values = []
    farmer_values = []
    for r in unique_rows:
        rsbsa = r[IDX["rsbsa"]]
        last = title_case_name(r[IDX["last"]])
        first = title_case_name(r[IDX["first"]])
        middle = title_case_name(r[IDX["middle"]])
        full_name = combine_name(last, first, middle)
        barangay = BARANGAY_MAP[r[IDX["barangay"]]]
        gender = SEX_MAP[r[IDX["sex"]]]
        birthdate = r[IDX["birthdate"]]
        farm_area = r[IDX["farm_area"]]

        household_id = deterministic_uuid(f"household:{rsbsa}")
        farmer_id = deterministic_uuid(f"farmer:{rsbsa}")

        display_name = f"Household — {full_name}"

        household_values.append(
            f"  ({sql_escape(household_id)}, {sql_escape(barangay)}, "
            f"{sql_escape(display_name)}, {fmt_float(farm_area)})"
        )
        farmer_values.append(
            f"  ({sql_escape(farmer_id)}, {sql_escape(full_name)}, "
            f"{sql_escape(gender)}, {sql_escape(barangay)}, "
            f"{sql_escape(household_id)}, TRUE, "
            f"{sql_escape(rsbsa)}, {fmt_date(birthdate)})"
        )

    out_lines.append(",\n".join(household_values))
    out_lines.append("ON CONFLICT (id) DO NOTHING;")
    out_lines.append("")
    out_lines.append("-- ─────────────────────────────────────────────────────────────────────")
    out_lines.append(f"-- Farmers ({len(unique_rows)} rows, each as head of their household)")
    out_lines.append("-- ─────────────────────────────────────────────────────────────────────")
    out_lines.append("")
    out_lines.append("INSERT INTO public.farmers (")
    out_lines.append("  id, name, gender, barangay, household_id, is_household_head, rsbsa_number, birth_date")
    out_lines.append(") VALUES")
    out_lines.append(",\n".join(farmer_values))
    out_lines.append("ON CONFLICT (id) DO NOTHING;")
    out_lines.append("")
    out_lines.append("COMMIT;")
    out_lines.append("")
    out_lines.append("-- =========================================================================")
    out_lines.append("-- Verification")
    out_lines.append("-- =========================================================================")
    out_lines.append("--")
    out_lines.append("--   SELECT 'households' AS t, count(*) FROM public.households")
    out_lines.append("--   UNION ALL SELECT 'farmers', count(*) FROM public.farmers;")
    out_lines.append(f"--   -- expect: households = farmers = {len(unique_rows)}")
    out_lines.append("--")
    out_lines.append("--   SELECT barangay, count(*) FROM public.farmers GROUP BY barangay ORDER BY barangay;")
    out_lines.append("--   -- expect, post-import:")
    for bgy, count in sorted(barangay_counts.items()):
        out_lines.append(f"--     {bgy:<12s} {count}")
    out_lines.append("--")
    out_lines.append("--   -- Spot-check a known farmer:")
    sample = unique_rows[0]
    sample_full = combine_name(
        title_case_name(sample[IDX["last"]]),
        title_case_name(sample[IDX["first"]]),
        title_case_name(sample[IDX["middle"]]),
    )
    out_lines.append(f"--   SELECT id, name, barangay, rsbsa_number, birth_date FROM public.farmers")
    out_lines.append(f"--   WHERE rsbsa_number = {sql_escape(sample[IDX['rsbsa']])};")
    out_lines.append(f"--   -- expect: name='{sample_full}', barangay='{BARANGAY_MAP[sample[IDX['barangay']]]}'")

    Path(sql_path).write_text("\n".join(out_lines) + "\n", encoding="utf-8")
    print(f"[ok] Wrote {sql_path}")
    print(f"  Unique farmers: {len(unique_rows)}")
    print(f"  Skipped duplicates: {skipped_dupes}")
    if skipped_unmapped_barangay:
        print(f"  Skipped unknown barangays: {Counter(skipped_unmapped_barangay).most_common(5)}")
    if skipped_unmapped_sex:
        print(f"  Skipped unknown sex values: {Counter(skipped_unmapped_sex).most_common(5)}")
    print()
    print("Per-barangay counts:")
    for bgy, count in sorted(barangay_counts.items()):
        print(f"  {bgy:<12s} {count}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    sys.exit(main(sys.argv[1], sys.argv[2]))
